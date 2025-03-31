import os
import openpyxl
import tkinter as tk
from tkinter import filedialog
import re

def extract_date_from_filename(filename):
    match = re.search(r'(\d{2})(\d{2})(\d{4})', filename)
    if match:
        return f"{match.group(1)}/{match.group(2)}/{match.group(3)}"
    return "Unknown Date"

def search_value_in_xlsx(folder_path, search_value):
    target_sheets = {"CHM0004", "CMP1049", "CMP2026", "CMP3740", "CMP9786", "EGR1018", "EGR2013", "EGR3043", "EGR9040",
                     "MTH1008", "MTH2009", "MTH3015", "PHY1005", "PHY2005", "PHY3010"}

    occurrences = []
    
    for file in os.listdir(folder_path):
        if file.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file)
            file_date = extract_date_from_filename(file)
            
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                
                for sheet_name in target_sheets:
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        for row in sheet.iter_rows(values_only=True):
                            if search_value in map(str, row):
                                occurrences.append((file, sheet_name, file_date))
                                break
                        
            except Exception as e:
                print(f"Error reading {file}: {e}")
    
    if not occurrences:
        print("No .xlsx files found in the folder or no matches found.")
        return False

    occurrences.sort(key=lambda x: tuple(map(int, x[2].split('/')[::-1])), reverse=True)
    
    output_file = os.path.join(folder_path, "search_results_test_export.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Search Results"
    sheet.append(["File", "Sheet", "File Date"])

    for file, sheet_name, file_date in occurrences:
        sheet.append([file, sheet_name, file_date])

    workbook.save(output_file)
    print(f"Results saved to {output_file}")

    return True

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    folder_path = filedialog.askdirectory(title="Select Folder Containing XLSX Files")
    
    if not folder_path:
        print("No folder selected. Exiting...")
    else:
        search_value = input("Enter the value to search for: ")  # Allow string input
        
        result = search_value_in_xlsx(folder_path, search_value)
        if result:
            print("The search results have been saved to an Excel file.")
        else:
            print("The value was NOT found in the specified sheets of the .xlsx files.")
